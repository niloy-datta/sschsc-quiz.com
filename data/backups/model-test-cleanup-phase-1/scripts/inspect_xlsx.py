import openpyxl
import os

wb_path = "HSC_Physics_2nd_Paper_AI_MCQ_Sets.xlsx"
if not os.path.exists(wb_path):
    print("File not found")
    exit()

wb = openpyxl.load_workbook(wb_path, read_only=True)
print("Sheet names:", wb.sheetnames)

sheet = wb.active
print("Active sheet title:", sheet.title)

for idx, row in enumerate(sheet.iter_rows(values_only=True)):
    if idx > 20:
        break
    print(f"Row {idx}: {row}")
